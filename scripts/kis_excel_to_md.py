"""
KIS OpenAPI 전체 문서 엑셀 → 마크다운 변환기.

엑셀 구조:
- "API 목록" 시트: 337개 API의 메타정보 인덱스 (메뉴 위치, API ID, TR_ID, URL, Method)
- 개별 API 시트 337개: 각 API의 Request/Response 명세 + Example

출력:
- docs/kis/00_INDEX.md: 전체 API 인덱스 (카테고리별 목차 + 전체 표)
- docs/kis/01_OAUTH.md ~ 22_BOND_REALTIME.md: 카테고리별 API 명세 묶음
"""
import openpyxl
import re
from collections import defaultdict
from pathlib import Path

XLSX = "/Users/koscom/Projects/stock-manager/docs/한국투자증권_오픈API_전체문서_20260508_030000.xlsx"
OUT_DIR = Path("/Users/koscom/Projects/stock-manager/docs/kis")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 카테고리 → 파일명 매핑 (순서 보존을 위해 list)
CATEGORY_FILES = [
    ("OAuth인증",                      "01_OAUTH.md",                "OAuth 인증"),
    ("[국내주식] 주문/계좌",            "02_KR_STOCK_ORDER.md",       "국내주식 주문/계좌"),
    ("[국내주식] 기본시세",             "03_KR_STOCK_QUOTE.md",       "국내주식 기본시세"),
    ("[국내주식] ELW 시세",             "04_KR_ELW.md",               "국내주식 ELW 시세"),
    ("[국내주식] 업종/기타",            "05_KR_SECTOR_ETC.md",        "국내주식 업종/기타"),
    ("[국내주식] 종목정보",             "06_KR_STOCK_INFO.md",        "국내주식 종목정보"),
    ("[국내주식] 시세분석",             "07_KR_MARKET_ANALYSIS.md",   "국내주식 시세분석"),
    ("[국내주식] 순위분석",             "08_KR_RANK.md",              "국내주식 순위분석"),
    ("[국내주식] 실시간시세",           "09_KR_REALTIME.md",          "국내주식 실시간시세"),
    ("[국내선물옵션] 주문/계좌",        "10_KR_FUTURES_ORDER.md",     "국내선물옵션 주문/계좌"),
    ("[국내선물옵션] 기본시세",         "11_KR_FUTURES_QUOTE.md",     "국내선물옵션 기본시세"),
    ("[국내선물옵션] 실시간시세",       "12_KR_FUTURES_REALTIME.md",  "국내선물옵션 실시간시세"),
    ("[해외주식] 주문/계좌",            "13_US_STOCK_ORDER.md",       "해외주식 주문/계좌"),
    ("[해외주식] 기본시세",             "14_US_STOCK_QUOTE.md",       "해외주식 기본시세"),
    ("[해외주식] 시세분석",             "15_US_STOCK_ANALYSIS.md",    "해외주식 시세분석"),
    ("[해외주식] 실시간시세",           "16_US_STOCK_REALTIME.md",    "해외주식 실시간시세"),
    ("[해외선물옵션] 주문/계좌",        "17_US_FUTURES_ORDER.md",     "해외선물옵션 주문/계좌"),
    ("[해외선물옵션] 기본시세",         "18_US_FUTURES_QUOTE.md",     "해외선물옵션 기본시세"),
    ("[해외선물옵션]실시간시세",        "19_US_FUTURES_REALTIME.md",  "해외선물옵션 실시간시세"),
    ("[장내채권] 주문/계좌",            "20_BOND_ORDER.md",           "장내채권 주문/계좌"),
    ("[장내채권] 기본시세",             "21_BOND_QUOTE.md",           "장내채권 기본시세"),
    ("[장내채권] 실시간시세",           "22_BOND_REALTIME.md",        "장내채권 실시간시세"),
]


def md_escape_pipe(s):
    """마크다운 표 셀 안에서 | 를 \\| 로 escape, 줄바꿈은 <br> 로."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("|", "\\|")
    s = s.replace("\n", "<br>")
    return s.strip()


def cell_str(v):
    if v is None:
        return ""
    return str(v).strip()


def make_anchor(text):
    """GitHub 마크다운 앵커 생성."""
    text = re.sub(r"[\(\)\[\]\{\}.,/]", "", str(text or ""))
    text = text.lower().replace(" ", "-")
    return text


def parse_api_sheet(ws):
    """API 시트 1개를 dict 로 파싱.

    엑셀 행 패턴(가변):
      R1: 시트제목(=API명)
      R2~: 메타정보 (label, value) 형태 — '기본정보', '개요', 'Layout', 'Example' 같은 섹션 헤더 행이 끼어듦
      Layout 섹션: 7열 표 (구분, Element, 한글명, Type, Required, Length, Description)
      Example 섹션: 2열 (label, code)

    Returns:
      dict {
        'title', 'meta': {label: value}, 'overview': str,
        'layout_rows': [(group, element, ko, type, req, length, desc), ...],
        'examples': [(label, code), ...],
        'extra': [(label, value)] — 위에 안 잡힌 metadata 행
      }
    """
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([cell_str(ws.cell(row=r, column=c).value) for c in range(1, 8)])

    out = {"title": rows[0][0] if rows else "", "meta": {}, "overview": "",
           "layout_rows": [], "examples": [], "extra": []}

    section = None  # None | 'meta' | 'overview' | 'layout' | 'example'
    for i, row in enumerate(rows):
        a, b = row[0], row[1]
        # 시트 제목
        if i == 0:
            continue
        # 섹션 헤더 (B열 비어있고 A열만 있는 헤더)
        if a in ("기본정보",) and not b:
            section = "meta"
            continue
        if a == "개요" and not b:
            section = "overview"
            continue
        if a == "Layout" and not b:
            section = "layout"
            continue
        if a == "Example" and not b:
            section = "example"
            continue

        # 빈 행 스킵
        if not any(c for c in row):
            continue

        # 처음 메타 블록: '기본정보' 헤더가 나오기 전까지 (API 통신방식, 메뉴 위치, API 명, API ID, 실전 TR_ID, 모의 TR_ID)
        if section is None:
            if a:
                out["meta"][a] = b
            continue

        if section == "meta":
            # HTTP Method, 실전 Domain, 모의 Domain, URL 명
            if a:
                out["meta"][a] = b
            continue

        if section == "overview":
            # '개요' label 다시 한번 + 본문
            if a == "개요" and b:
                out["overview"] = b
            elif a and not b:
                # rare — 개요 헤더 두 번째 라벨
                pass
            elif b:
                # value 만 있는 케이스
                out["overview"] += "\n" + b
            continue

        if section == "layout":
            # 첫 행은 "구분 / Element / 한글명 / Type / Required / Length / Description" — 헤더 자체
            if a == "구분" and b == "Element":
                continue
            # 데이터 행 (group, element, ko, type, req, len, desc)
            out["layout_rows"].append(tuple(row[:7]))
            continue

        if section == "example":
            # (label, code) — code 는 \r\n 포함 멀티라인
            if a and b:
                out["examples"].append((a, b))
            elif a and not b:
                out["examples"].append((a, ""))
            continue

    return out


def render_api_md(parsed, level=2):
    """파싱된 API 한 건을 마크다운으로 직렬화."""
    h = "#" * level
    out = []
    title = parsed["title"] or "(제목 없음)"
    out.append(f"{h} {title}")
    out.append("")

    # 메타 표
    meta = parsed["meta"]
    if meta:
        out.append(f"{'#' * (level + 1)} 기본 정보")
        out.append("")
        out.append("| 항목 | 값 |")
        out.append("| --- | --- |")
        # 권장 표시 순서
        order = [
            "API 통신방식", "메뉴 위치", "API 명", "API ID",
            "실전 TR_ID", "모의 TR_ID",
            "HTTP Method", "URL 명", "실전 Domain", "모의 Domain",
        ]
        seen = set()
        for k in order:
            if k in meta:
                out.append(f"| {md_escape_pipe(k)} | {md_escape_pipe(meta[k])} |")
                seen.add(k)
        for k, v in meta.items():
            if k not in seen:
                out.append(f"| {md_escape_pipe(k)} | {md_escape_pipe(v)} |")
        out.append("")

    # 개요
    if parsed["overview"]:
        out.append(f"{'#' * (level + 1)} 개요")
        out.append("")
        # 개요 본문은 줄바꿈 보존 (단순 paragraph)
        text = parsed["overview"].replace("\r\n", "\n").replace("\r", "\n").strip()
        out.append(text)
        out.append("")

    # Layout 표
    if parsed["layout_rows"]:
        out.append(f"{'#' * (level + 1)} Layout")
        out.append("")
        # 구분(group) 별로 분리해서 표 4개로 보여주는 것이 더 가독성 좋다
        by_group = []
        cur_group = None
        cur_rows = []
        for row in parsed["layout_rows"]:
            grp = row[0] or cur_group
            if grp != cur_group and grp:
                if cur_rows:
                    by_group.append((cur_group, cur_rows))
                cur_group = grp
                cur_rows = [row]
            else:
                cur_rows.append(row)
        if cur_rows:
            by_group.append((cur_group, cur_rows))

        for grp, group_rows in by_group:
            if grp:
                out.append(f"**{grp}**")
                out.append("")
            out.append("| Element | 한글명 | Type | Required | Length | Description |")
            out.append("| --- | --- | --- | --- | --- | --- |")
            for r in group_rows:
                # r = (group, element, ko, type, req, len, desc)
                _, element, ko, typ, req, length, desc = r
                if not any([element, ko, typ, req, length, desc]):
                    continue
                out.append(
                    f"| {md_escape_pipe(element)} | {md_escape_pipe(ko)} | "
                    f"{md_escape_pipe(typ)} | {md_escape_pipe(req)} | "
                    f"{md_escape_pipe(length)} | {md_escape_pipe(desc)} |"
                )
            out.append("")

    # Example
    if parsed["examples"]:
        out.append(f"{'#' * (level + 1)} Example")
        out.append("")
        for label, code in parsed["examples"]:
            out.append(f"**{label}**")
            out.append("")
            out.append("```")
            text = code.replace("\r\n", "\n").replace("\r", "\n")
            out.append(text)
            out.append("```")
            out.append("")

    return "\n".join(out)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    api_list = wb["API 목록"]

    # 메뉴 위치 → [(순번, API명, API ID, 실전 TR_ID, 모의 TR_ID, Method, URL, 통신방식)]
    rows_by_cat = defaultdict(list)
    full_rows = []
    for r in range(2, api_list.max_row + 1):
        cells = [cell_str(api_list.cell(row=r, column=c).value) for c in range(1, 12)]
        # [순번, 통신방식, 메뉴 위치, API 명, API ID, 실전 TR_ID, 모의 TR_ID, HTTP Method, URL 명, 실전 Domain, 모의 Domain]
        if not cells[0]:
            continue
        rows_by_cat[cells[2]].append(cells)
        full_rows.append(cells)

    # ── 카테고리별 md 작성 ───────────────────────
    written = []
    for cat, fname, ko_title in CATEGORY_FILES:
        rows = rows_by_cat.get(cat, [])
        if not rows:
            continue
        out = []
        out.append(f"# {ko_title}")
        out.append("")
        out.append(f"**카테고리 코드**: `{cat}`  ")
        out.append(f"**API 수**: {len(rows)}개")
        out.append("")
        out.append("> 출처: 한국투자증권 OpenAPI 전체 문서 (2026-05-08).  ")
        out.append("> 자동 변환: `scripts/kis_excel_to_md.py`")
        out.append("")
        out.append("---")
        out.append("")
        out.append("## 목차")
        out.append("")
        for cells in rows:
            api_name = cells[3]
            anchor = make_anchor(api_name)
            tr = cells[5] or "—"
            out.append(f"- [{api_name}](#{anchor}) — `{cells[7]}` `{cells[8]}` (실전 TR_ID: `{tr}`)")
        out.append("")
        out.append("---")
        out.append("")

        # 각 API 시트 본문
        for cells in rows:
            api_name = cells[3]
            if api_name not in wb.sheetnames:
                # sheet 이름이 다른 경우 fallback (괄호/공백 차이)
                # 일치하는 시트 시도
                candidates = [s for s in wb.sheetnames if s.replace(" ", "") == api_name.replace(" ", "")]
                if candidates:
                    sheet_name = candidates[0]
                else:
                    out.append(f"## {api_name}\n\n> ⚠️ 시트를 찾지 못했습니다.\n")
                    continue
            else:
                sheet_name = api_name
            ws = wb[sheet_name]
            parsed = parse_api_sheet(ws)
            # 시트 제목이 비어있을 수 있어 보강
            if not parsed["title"]:
                parsed["title"] = api_name
            # 카테고리 메타도 보강
            parsed["meta"].setdefault("순번", cells[0])
            parsed["meta"].setdefault("API 통신방식", cells[1])
            parsed["meta"].setdefault("메뉴 위치", cells[2])
            parsed["meta"].setdefault("API 명", cells[3])
            parsed["meta"].setdefault("API ID", cells[4])
            parsed["meta"].setdefault("실전 TR_ID", cells[5])
            parsed["meta"].setdefault("모의 TR_ID", cells[6])
            parsed["meta"].setdefault("HTTP Method", cells[7])
            parsed["meta"].setdefault("URL 명", cells[8])
            parsed["meta"].setdefault("실전 Domain", cells[9])
            parsed["meta"].setdefault("모의 Domain", cells[10])

            md = render_api_md(parsed, level=2)
            out.append(md)
            out.append("---")
            out.append("")

        path = OUT_DIR / fname
        path.write_text("\n".join(out), encoding="utf-8")
        written.append((fname, ko_title, cat, len(rows)))
        print(f"[WROTE] {fname} — {len(rows)}개 API")

    # ── 00_INDEX.md ────────────────────────────
    out = []
    out.append("# KIS OpenAPI 전체 인덱스")
    out.append("")
    out.append("> 출처: `한국투자증권_오픈API_전체문서_20260508_030000.xlsx` (2026-05-08 기준).  ")
    out.append("> 자동 변환: `scripts/kis_excel_to_md.py`. 본 디렉토리는 코드/리뷰용 참조 자료이며, 1:1 사본이 아님(표 형식 정리됨).")
    out.append("")
    out.append("## 카테고리 목차")
    out.append("")
    out.append("| 파일 | 카테고리 | API 수 |")
    out.append("| --- | --- | --- |")
    for fname, ko_title, cat, cnt in written:
        out.append(f"| [{fname}](./{fname}) | {ko_title} (`{cat}`) | {cnt} |")
    out.append("")
    total = sum(cnt for *_ , cnt in written)
    out.append(f"**총 API 수**: {total}개  ")
    out.append("")
    out.append("---")
    out.append("")
    out.append("## 전체 API 표 (메뉴 위치별 정렬)")
    out.append("")
    out.append("| # | 카테고리 | API 명 | API ID | 실전 TR_ID | 모의 TR_ID | Method | URL |")
    out.append("| --- | --- | --- | --- | --- | --- | --- | --- |")
    for cat, _fname, _ko in CATEGORY_FILES:
        for cells in rows_by_cat.get(cat, []):
            # cells: [순번, 통신방식, 메뉴 위치, API 명, API ID, 실전 TR_ID, 모의 TR_ID, HTTP Method, URL 명, 실전 Domain, 모의 Domain]
            out.append(
                f"| {md_escape_pipe(cells[0])} | {md_escape_pipe(cells[2])} | "
                f"{md_escape_pipe(cells[3])} | {md_escape_pipe(cells[4])} | "
                f"{md_escape_pipe(cells[5])} | {md_escape_pipe(cells[6])} | "
                f"{md_escape_pipe(cells[1] + ' ' + cells[7]).strip()} | "
                f"`{md_escape_pipe(cells[8])}` |"
            )
    out.append("")
    out.append("---")
    out.append("")
    out.append("## 도메인")
    out.append("")
    out.append("| 환경 | URL |")
    out.append("| --- | --- |")
    out.append("| 실전 | `https://openapi.koreainvestment.com:9443` |")
    out.append("| 모의 | `https://openapivts.koreainvestment.com:29443` |")
    out.append("")
    out.append("> 모의투자 미지원 API는 카테고리별 문서의 `모의 TR_ID` 필드에 명시됨.")
    out.append("")
    (OUT_DIR / "00_INDEX.md").write_text("\n".join(out), encoding="utf-8")
    print(f"[WROTE] 00_INDEX.md — total {total}개 API 인덱스")


if __name__ == "__main__":
    main()
